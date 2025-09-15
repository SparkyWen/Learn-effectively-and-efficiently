
# -*- coding: utf-8 -*-
"""
Excel 表头批改助手（.xlsx）
-----------------------------------
需求：
  - 对一个文件夹中的所有 .xlsx 文件，只改第一行表头为指定的列名；其他内容不变。
  - 支持简单 GUI。
  - 支持并行（线程池/进程池）加速。
  - 列名可自定义。

实现要点：
  - 使用 openpyxl 直接修改第一行单元格，尽量保持格式、公式等其它内容不变。
  - 可选择处理所有 Sheet / 仅第一个 Sheet / 只处理指定名称的 Sheet。
  - 可选择输出到新目录（推荐，安全）或原地覆盖（可选备份）。
  - 表头写入策略可选：
      * 只覆盖已有列（不改变列数）【默认】
      * 覆盖已有列 + 多余列标题清空
      * 强制按目标列写满（可能超出现有列数）
  - 并行处理文件，日志可视化，支持取消。
"""
import os
import sys
import json
import time
import shutil
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed

# GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CFG_FILE = "excel_header_renamer.config.json"
TEMP_PREFIX = "~$"

# ---------- 工具函数 ----------
def log_exception(e: BaseException) -> str:
    return "".join(traceback.format_exception(type(e), e, e.__traceback__))

def ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def collect_xlsx(root: Path, recursive: bool) -> List[Path]:
    if recursive:
        files = [p for p in root.rglob("*.xlsx")]
    else:
        files = list(root.glob("*.xlsx"))
    files = [p for p in files if not p.name.startswith(TEMP_PREFIX)]
    files.sort()
    return files

def parse_header_list(s: str) -> List[str]:
    """把用户输入的一行或多行文本解析为列名列表，支持逗号/中文逗号/分号/换行分隔"""
    if not s:
        return []
    for ch in ["，", "；", ";", "\n", "\r"]:
        s = s.replace(ch, ",")
    parts = [x.strip() for x in s.split(",")]
    parts = [x for x in parts if x != ""]
    return parts

def trim_trailing_nones(values: List[Any]) -> List[Any]:
    i = len(values) - 1
    while i >= 0 and (values[i] is None or (isinstance(values[i], str) and values[i].strip() == "")):
        i -= 1
    return values[: i + 1]

def infer_ncols(ws) -> int:
    """尽量准确地根据第一行推断实际列数（去掉尾部空单元格）"""
    try:
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row1 = list(row1) if row1 is not None else []
        trimmed = trim_trailing_nones(row1)
        if trimmed:
            return len(trimmed)
    except Exception:
        pass
    # 回退：使用 ws.max_column（可能偏大）
    return ws.max_column or 0

# ---------- 并行 Worker ----------
@dataclass
class TaskOptions:
    headers: List[str]
    sheet_mode: str          # "all" | "first" | "byname"
    sheet_names: List[str]   # 当 sheet_mode=="byname" 时生效
    write_mode: str          # "cover_exist" | "cover_and_blank" | "force_target"
    inplace: bool
    backup: bool
    out_dir: Optional[str]   # 当 inplace=False 时，输出目录
    keep_tree: bool          # 保留子目录结构（递归时）
    suffix: str              # 输出文件名后缀（不含扩展名）

def rename_headers_worker(file_path: str, opts: TaskOptions) -> Dict[str, Any]:
    """处理单个文件：返回 {ok, file, out, sheets_done, error}"""
    fpath = Path(file_path)
    try:
        wb = load_workbook(fpath)
    except Exception as e:
        return {"ok": False, "file": str(fpath), "out": None, "sheets_done": 0, "error": f"读取失败: {e}"}

    # Sheet 过滤
    all_sheetnames = wb.sheetnames[:]
    target_sheets: List[str] = []
    if opts.sheet_mode == "all":
        target_sheets = all_sheetnames
    elif opts.sheet_mode == "first":
        target_sheets = all_sheetnames[:1]
    else:  # byname
        allow = set([s.lower() for s in opts.sheet_names])
        target_sheets = [s for s in all_sheetnames if s.lower() in allow]

    renamed = 0
    for sname in target_sheets:
        try:
            ws = wb[sname]
            ncols = infer_ncols(ws)
            if ncols <= 0 and ws.max_column > 0:
                ncols = ws.max_column

            # 写入策略
            if opts.write_mode == "cover_exist":
                # 仅覆盖已有列头，不改变列数
                m = min(ncols, len(opts.headers))
                for i in range(m):
                    ws.cell(row=1, column=i + 1, value=opts.headers[i])
                renamed += 1
            elif opts.write_mode == "cover_and_blank":
                # 覆盖已有列，且对超过目标的列头置空
                for i in range(ncols):
                    val = opts.headers[i] if i < len(opts.headers) else ""
                    ws.cell(row=1, column=i + 1, value=val)
                renamed += 1
            else:  # force_target
                # 强制写满目标列（可比当前列数更多）
                for i in range(len(opts.headers)):
                    ws.cell(row=1, column=i + 1, value=opts.headers[i])
                renamed += 1
        except Exception as e:
            # 单个 sheet 失败不影响其他 sheet
            pass

    # 输出路径
    if opts.inplace:
        # 备份
        if opts.backup:
            try:
                bak = fpath.with_suffix(fpath.suffix + f".bak.{int(time.time())}")
                shutil.copy2(fpath, bak)
            except Exception:
                pass
        out_path = fpath
    else:
        out_root = Path(opts.out_dir or (fpath.parent / "renamed_headers"))
        if opts.keep_tree:
            # 保留原相对结构：由调用方提供 root 以外的相对路径，这里无法直接获知；
            # 简化处理：只在与源同级创建 out_root，不再额外拼子路径（UI 已提示）。
            pass
        # 文件名加后缀
        new_name = fpath.stem + (opts.suffix or "") + fpath.suffix
        out_path = out_root / new_name
        ensure_parent(out_path)

    try:
        wb.save(out_path)
    except Exception as e:
        return {"ok": False, "file": str(fpath), "out": str(out_path), "sheets_done": renamed, "error": f"保存失败: {e}"}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {"ok": True, "file": str(fpath), "out": str(out_path), "sheets_done": renamed, "error": None}

# ---------- GUI 应用 ----------
@dataclass
class Options:
    in_dir: str
    recursive: bool
    headers: List[str]
    sheet_mode: str
    sheet_names: List[str]
    write_mode: str
    inplace: bool
    backup: bool
    out_dir: str
    keep_tree: bool
    suffix: str
    parallel: bool
    kind: str         # "process" | "thread"
    max_workers: int

class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Excel 表头批改助手（.xlsx）- 并行加速")
        self.geometry("900x680")
        self.minsize(900, 680)
        self._working = False
        self._build_ui()
        self._load_cfg()

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        # 顶部：目录与输出
        top = ttk.Frame(self)
        top.pack(fill="x", **pad)
        ttk.Label(top, text="输入文件夹：").grid(row=0, column=0, sticky="e")
        self.dir_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.dir_var, width=70).grid(row=0, column=1, sticky="we", padx=(0,6))
        ttk.Button(top, text="浏览...", command=self._choose_dir).grid(row=0, column=2, sticky="w")
        top.columnconfigure(1, weight=1)

        # 目标列名
        frm_hdr = ttk.LabelFrame(self, text="目标列名（自定义，逗号/换行分隔）")
        frm_hdr.pack(fill="x", **pad)
        self.headers_text = tk.Text(frm_hdr, height=4, wrap="word")
        self.headers_text.pack(fill="x", expand=True, padx=8, pady=6)
        self.headers_text.insert("1.0", "产品名称, 时间, 行业/应用领域, 官网链接, 一句话介绍（功能与地位）")

        # Sheet 处理方式
        frm_sheet = ttk.LabelFrame(self, text="工作表（Sheet）选择")
        frm_sheet.pack(fill="x", **pad)
        self.sheet_mode = tk.StringVar(value="first")
        ttk.Radiobutton(frm_sheet, text="仅第一个 Sheet", variable=self.sheet_mode, value="first").grid(row=0, column=0, sticky="w", padx=8, pady=2)
        ttk.Radiobutton(frm_sheet, text="所有 Sheet", variable=self.sheet_mode, value="all").grid(row=0, column=1, sticky="w", padx=8, pady=2)
        ttk.Radiobutton(frm_sheet, text="只处理指定名称", variable=self.sheet_mode, value="byname").grid(row=0, column=2, sticky="w", padx=8, pady=2)
        ttk.Label(frm_sheet, text="指定 Sheet 名（逗号/换行分隔，可留空）：").grid(row=1, column=0, sticky="e", padx=(8,2))
        self.sheet_names_var = tk.StringVar(value="")
        ttk.Entry(frm_sheet, textvariable=self.sheet_names_var, width=60).grid(row=1, column=1, columnspan=2, sticky="w")

        # 表头写入策略
        frm_mode = ttk.LabelFrame(self, text="表头写入策略")
        frm_mode.pack(fill="x", **pad)
        self.write_mode = tk.StringVar(value="cover_exist")
        ttk.Radiobutton(frm_mode, text="只覆盖已有列（不改变列数）", variable=self.write_mode, value="cover_exist").grid(row=0, column=0, sticky="w", padx=8, pady=2)
        ttk.Radiobutton(frm_mode, text="覆盖已有列 + 多余列标题清空", variable=self.write_mode, value="cover_and_blank").grid(row=0, column=1, sticky="w", padx=8, pady=2)
        ttk.Radiobutton(frm_mode, text="强制按目标列写满（可能超出当前列数）", variable=self.write_mode, value="force_target").grid(row=0, column=2, sticky="w", padx=8, pady=2)

        # 输出选项
        frm_out = ttk.LabelFrame(self, text="输出方式")
        frm_out.pack(fill="x", **pad)
        self.inplace_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_out, text="原地覆盖（谨慎，建议勾选“生成备份”）", variable=self.inplace_var, command=self._toggle_out_fields).grid(row=0, column=0, sticky="w", padx=8)
        self.backup_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_out, text="为原文件生成 .bak 备份", variable=self.backup_var).grid(row=0, column=1, sticky="w", padx=8)

        ttk.Label(frm_out, text="输出文件夹（不覆盖时生效）：").grid(row=1, column=0, sticky="e")
        self.out_dir_var = tk.StringVar(value="")
        ttk.Entry(frm_out, textvariable=self.out_dir_var, width=60).grid(row=1, column=1, sticky="we", padx=(0,6))
        ttk.Button(frm_out, text="选择...", command=self._choose_out_dir).grid(row=1, column=2, sticky="w")
        frm_out.columnconfigure(1, weight=1)

        ttk.Label(frm_out, text="输出文件名后缀：").grid(row=2, column=0, sticky="e")
        self.suffix_var = tk.StringVar(value="_renamed")
        ttk.Entry(frm_out, textvariable=self.suffix_var, width=20).grid(row=2, column=1, sticky="w")

        # 其它选项
        frm_misc = ttk.LabelFrame(self, text="其它")
        frm_misc.pack(fill="x", **pad)
        self.recursive_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_misc, text="递归子目录", variable=self.recursive_var).grid(row=0, column=0, sticky="w", padx=8)

        # 并行
        frm_par = ttk.LabelFrame(self, text="并行加速")
        frm_par.pack(fill="x", **pad)
        self.parallel_var = tk.BooleanVar(value=True)
        self.kind_var = tk.StringVar(value="process")
        self.workers_var = tk.IntVar(value=max(2, (os.cpu_count() or 4)//2))
        ttk.Checkbutton(frm_par, text="启用并行", variable=self.parallel_var).grid(row=0, column=0, sticky="w", padx=8)
        ttk.Label(frm_par, text="方式：").grid(row=0, column=1, sticky="e")
        ttk.Combobox(frm_par, textvariable=self.kind_var, values=["process", "thread"], width=10, state="readonly").grid(row=0, column=2, sticky="w", padx=(2,8))
        ttk.Label(frm_par, text="最大并发：").grid(row=0, column=3, sticky="e")
        ttk.Spinbox(frm_par, from_=1, to=max(64, (os.cpu_count() or 8)), textvariable=self.workers_var, width=6).grid(row=0, column=4, sticky="w", padx=(2,8))

        # 运行区
        run = ttk.Frame(self)
        run.pack(fill="x", **pad)
        self.btn_start = ttk.Button(run, text="开始批改", command=self._on_start)
        self.btn_start.pack(side="left")
        self.btn_stop = ttk.Button(run, text="取消任务", command=self._on_stop, state="disabled")
        self.btn_stop.pack(side="left", padx=(8,0))
        self.prog = ttk.Progressbar(run, mode="determinate")
        self.prog.pack(side="right", fill="x", expand=True)

        # 日志
        frm_log = ttk.LabelFrame(self, text="日志")
        frm_log.pack(fill="both", expand=True, **pad)
        self.txt = tk.Text(frm_log, height=18, wrap="word")
        sb = ttk.Scrollbar(frm_log, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        self.txt.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def _toggle_out_fields(self):
        inplace = self.inplace_var.get()
        state = "disabled" if inplace else "normal"
        # 输出目录与后缀仅在 not inplace 时可编辑（后缀也可以在 inplace 下使用，但容易造成误解，默认禁用）
        for w in []:
            w.configure(state=state)

    def _choose_dir(self):
        d = filedialog.askdirectory(title="选择包含 .xlsx 的文件夹")
        if d:
            self.dir_var.set(d)

    def _choose_out_dir(self):
        d = filedialog.askdirectory(title="选择输出文件夹")
        if d:
            self.out_dir_var.set(d)

    # ---------- 配置 ----------
    def _load_cfg(self):
        try:
            cfgp = Path(CFG_FILE)
            if cfgp.exists():
                cfg = json.loads(cfgp.read_text(encoding="utf-8"))
                self.dir_var.set(cfg.get("in_dir", ""))
                self.out_dir_var.set(cfg.get("out_dir", ""))
                self.headers_text.delete("1.0", "end")
                self.headers_text.insert("1.0", cfg.get("headers_text", "产品名称, 时间, 行业/应用领域, 官网链接, 一句话介绍（功能与地位）"))
        except Exception:
            pass

    def _save_cfg(self):
        try:
            cfg = {
                "in_dir": self.dir_var.get().strip(),
                "out_dir": self.out_dir_var.get().strip(),
                "headers_text": self.headers_text.get("1.0", "end").strip()
            }
            Path(CFG_FILE).write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    # ---------- 日志 ----------
    def log(self, msg: str):
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")
        self.update_idletasks()

    def set_prog(self, value: float, maximum: Optional[float] = None):
        if maximum is not None:
            self.prog.configure(maximum=maximum)
        self.prog.configure(value=value)
        self.update_idletasks()

    # ---------- 运行 ----------
    def _on_stop(self):
        self._working = False
        self.btn_stop.configure(state="disabled")
        self.log("收到取消请求，将在当前批次结束后停止。")

    def _on_start(self):
        if self._working:
            return
        in_dir = self.dir_var.get().strip()
        if not in_dir:
            messagebox.showwarning("提示", "请选择输入文件夹")
            return
        root = Path(in_dir)
        if not root.exists() or not root.is_dir():
            messagebox.showerror("错误", "输入路径不存在或不是文件夹")
            return

        headers = parse_header_list(self.headers_text.get("1.0", "end").strip())
        if not headers:
            messagebox.showwarning("提示", "请填写目标列名（至少一个）")
            return

        sheet_mode = self.sheet_mode.get()
        sheet_names = parse_header_list(self.sheet_names_var.get().strip())

        write_mode = self.write_mode.get()
        inplace = self.inplace_var.get()
        backup = self.backup_var.get()
        out_dir = self.out_dir_var.get().strip() or str(root / "renamed_headers")
        suffix = self.suffix_var.get().strip()
        recursive = self.recursive_var.get()

        parallel = self.parallel_var.get()
        kind = self.kind_var.get()
        max_workers = max(1, int(self.workers_var.get() or 1))

        self._save_cfg()

        # 收集文件
        files = collect_xlsx(root, recursive=recursive)
        if not files:
            self.log("未在该目录下找到 .xlsx 文件。")
            messagebox.showinfo("提示", "未找到 .xlsx 文件")
            return

        if not inplace and not out_dir:
            out_dir = str(root / "renamed_headers")
        if not inplace:
            Path(out_dir).mkdir(parents=True, exist_ok=True)

        opts = Options(
            in_dir=str(root),
            recursive=recursive,
            headers=headers,
            sheet_mode=sheet_mode,
            sheet_names=sheet_names,
            write_mode=write_mode,
            inplace=inplace,
            backup=backup,
            out_dir=out_dir,
            keep_tree=False,
            suffix=suffix,
            parallel=parallel,
            kind=kind,
            max_workers=max_workers,
        )

        # 启动任务
        self._working = True
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.txt.delete("1.0", "end")
        self.set_prog(0, maximum=len(files))

        self.after(100, lambda: self._run(files, opts))

    def _run(self, files: List[Path], opts: Options):
        try:
            self.log(f"发现 {len(files)} 个 .xlsx 文件。开始处理...")
            t0 = time.time()

            task_opts = TaskOptions(
                headers=opts.headers,
                sheet_mode=opts.sheet_mode,
                sheet_names=opts.sheet_names,
                write_mode=opts.write_mode,
                inplace=opts.inplace,
                backup=opts.backup,
                out_dir=opts.out_dir,
                keep_tree=opts.keep_tree,
                suffix=opts.suffix,
            )

            done = 0
            ok_cnt = 0
            if opts.parallel:
                ex_cls = ProcessPoolExecutor if opts.kind == "process" else ThreadPoolExecutor
                self.log(f"并行模式：{opts.kind}，并发数：{opts.max_workers}")
                with ex_cls(max_workers=opts.max_workers) as ex:
                    fut_map = {}
                    for f in files:
                        if not self._working:
                            break
                        fut = ex.submit(rename_headers_worker, str(f), task_opts)
                        fut_map[fut] = f
                    for fut in as_completed(fut_map):
                        if not self._working:
                            break
                        try:
                            res = fut.result()
                        except Exception as e:
                            self.log(f"[异常] {fut_map[fut].name}: {e}")
                            done += 1
                            self.set_prog(done)
                            continue
                        done += 1
                        if res.get("ok"):
                            ok_cnt += 1
                            self.log(f"[OK] {Path(res['file']).name} -> {Path(res['out']).name} （修改 {res['sheets_done']} 个Sheet）")
                        else:
                            self.log(f"[失败] {Path(res['file']).name}: {res.get('error')}")
                        self.set_prog(done)
            else:
                for f in files:
                    if not self._working:
                        break
                    res = rename_headers_worker(str(f), task_opts)
                    done += 1
                    if res.get("ok"):
                        ok_cnt += 1
                        self.log(f"[OK] {Path(res['file']).name} -> {Path(res['out']).name} （修改 {res['sheets_done']} 个Sheet）")
                    else:
                        self.log(f"[失败] {Path(res['file']).name}: {res.get('error')}")
                    self.set_prog(done)

            dt = time.time() - t0
            if not self._working:
                self.log("任务已取消。")
            else:
                self.log(f"✅ 完成：成功 {ok_cnt}/{len(files)} 个文件，用时 {dt:.2f}s")
                try:
                    if sys.platform.startswith("win"):
                        os.startfile(opts.out_dir if not opts.inplace else opts.in_dir)  # type: ignore
                except Exception:
                    pass
                messagebox.showinfo("完成", f"成功处理 {ok_cnt}/{len(files)} 个文件。")
        except Exception as e:
            self.log("发生异常：\n" + log_exception(e))
            messagebox.showerror("错误", str(e))
        finally:
            self._working = False
            self.btn_start.configure(state="normal")
            self.btn_stop.configure(state="disabled")

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    try:
        from multiprocessing import freeze_support
        freeze_support()
    except Exception:
        pass
    main()
